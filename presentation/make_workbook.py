"""Results workbook: formatted results table + column definitions + scenario notes."""
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent
OUT = Path(__file__).resolve().parent / "esaf_results_v8.xlsx"

runs = {k: pd.read_csv(BASE / d / "summary_all_airports.csv").set_index("country_iso3")
        for k, d in [("fixed", "results_v8_fixed"), ("m2030", "results_v8_2030"),
                     ("t30", "results_v8_2050t30"), ("t50", "results_v8_2050t50")]}
REF3 = ["SAU", "ARE", "MAR"]
eu = [c for c in runs["fixed"].index if c not in REF3]
order = runs["t50"].loc[eu, "delivered_cost_usd_per_tonne_saf"].sort_values().index.tolist() + REF3

STRAT = {"flex_fuel_store": "flexible + fuel tanks", "steady_h2_store": "steady + H2 cavern"}
CARB = {"market": "biogenic (purchased)", "dac": "DAC (on-site)"}

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
BAND = PatternFill("solid", fgColor="F2F2F2")
REFBAND = PatternFill("solid", fgColor="FBF3E4")
INPUT_FONT = Font(name=ARIAL, color="0000FF")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(bottom=thin)

wb = Workbook()

# ------------------------------------------------------------------ Results
ws = wb.active
ws.title = "Results"
ws.sheet_view.showGridLines = False

ws["A1"] = "Delivered cost of e-SAF at the national hub airport — model v8 results"
ws["A1"].font = Font(name=ARIAL, bold=True, size=14)
ws["A2"] = "One plant per country, cheapest feasible site; all monetary values in real USD-2024 unless marked EUR. See 'Column definitions' sheet before reading."
ws["A2"].font = Font(name=ARIAL, size=9, italic=True, color="595959")
ws["A3"] = "USD → EUR conversion (2024 ECB annual average):"
ws["A3"].font = Font(name=ARIAL, size=9)
ws["D3"] = 0.9239
ws["D3"].font = INPUT_FONT
ws["D3"].fill = PatternFill("solid", fgColor="FFFF00")
ws["D3"].number_format = "0.0000"
ws["E3"] = "editable input cell (blue on yellow); EUR columns recalculate from it"
ws["E3"].font = Font(name=ARIAL, size=8, italic=True, color="595959")

headers = [
    ("Country", 8), ("Country name", 15), ("Hub airport", 24), ("IATA", 7),
    ("Hub fuel 2024 (kt)", 11),
    ("e-SAF share 2030", 10), ("Plant size 2030 (kt/y)", 11), ("Cost 2030 (USD/t)", 12),
    ("Cost 2030 (EUR/t)", 12), ("Strategy 2030", 19), ("Carbon 2030", 19),
    ("Bio-CO2 price 2030 (USD/t CO2)", 12),
    ("e-SAF share 2050", 10), ("Plant size 2050 (kt/y)", 11),
    ("Cost 2050 @2030 tech (USD/t)", 13), ("Cost 2050 @2050 tech (USD/t)", 13),
    ("Cost 2050 @2050 tech (EUR/t)", 13), ("Strategy 2050", 19), ("Carbon 2050", 19),
    ("Reference 740 kt (USD/t)", 12), ("WACC (real)", 9), ("H2 storage geology", 11),
]
HR = 5
for j, (h, w) in enumerate(headers, start=1):
    c = ws.cell(row=HR, column=j, value=h)
    c.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=9)
    c.fill = HDR_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[HR].height = 42

r = HR
for iso3 in order:
    r += 1
    ref = iso3 in REF3
    f0, m30, t30, t50 = (runs[k].loc[iso3] for k in ["fixed", "m2030", "t30", "t50"])
    name = f0.get("country_name", iso3)
    vals = [
        iso3, name, f0["airport_name"], f0["airport_iata"],
        round(float(m30.get("airport_fuel_kt_2024") or 0), 0) or None,
    ]
    if ref:
        vals += [None, None, None, None, "no mandate — export case", "DAC (on-site)", None,
                 None, None, None, round(float(t50["delivered_cost_usd_per_tonne_saf"])), None,
                 STRAT.get(t50["operating_strategy"], ""), CARB.get(t50["carbon_source"], "")]
    else:
        vals += [
            float(m30["mandate_synthetic_share"]),
            round(float(m30["annual_saf_tonnes"]) / 1000, 1),
            round(float(m30["delivered_cost_usd_per_tonne_saf"])),
            None,  # EUR formula below
            STRAT.get(m30["operating_strategy"], ""), CARB.get(m30["carbon_source"], ""),
            round(float(m30["co2_delivered_cost_usd_per_t"])) if m30["carbon_source"] == "market" else None,
            float(t50["mandate_synthetic_share"]),
            round(float(t50["annual_saf_tonnes"]) / 1000, 1),
            round(float(t30["delivered_cost_usd_per_tonne_saf"])),
            round(float(t50["delivered_cost_usd_per_tonne_saf"])),
            None,  # EUR formula below
            STRAT.get(t50["operating_strategy"], ""), CARB.get(t50["carbon_source"], ""),
        ]
    vals += [round(float(f0["delivered_cost_usd_per_tonne_saf"])),
             float(f0["wacc_used"]), f0["h2_storage_geology"]]
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = Font(name=ARIAL, size=9)
        c.border = BORDER
        if (r - HR) % 2 == 0:
            c.fill = BAND
        if ref:
            c.fill = REFBAND
    if not ref:
        ws.cell(row=r, column=9, value=f"=H{r}*$D$3").font = Font(name=ARIAL, size=9)
        ws.cell(row=r, column=17, value=f"=P{r}*$D$3").font = Font(name=ARIAL, size=9)
    for col, fmt in [(5, "#,##0"), (6, "0.0%"), (7, "#,##0.0"), (8, "#,##0"), (9, "#,##0"),
                     (12, "#,##0"), (13, "0.0%"), (14, "#,##0.0"), (15, "#,##0"),
                     (16, "#,##0"), (17, "#,##0"), (20, "#,##0"), (21, "0.0%")]:
        ws.cell(row=r, column=col).number_format = fmt

last = r
ws.freeze_panes = f"E{HR+1}"
ws.auto_filter.ref = f"A{HR}:V{last}"
note = ws.cell(row=last + 2, column=1,
    value="Reference cases (amber rows): Saudi Arabia, UAE, Morocco carry no EU/UK mandate; shown at the fixed 740 kt/y reference size pending the export/import analysis. "
          "Sorted by 2050 cost at 2050 technology (EU/UK/CH), cheapest first.")
note.font = Font(name=ARIAL, size=8, italic=True, color="595959")

# --------------------------------------------------------- Cost breakdown
bd = pd.read_csv(Path(__file__).resolve().parent / "cost_breakdown_v8.csv")
bd.insert(len(bd.columns) - 3, "small_lines_subtotal",
          (bd["fuel_transport"] + bd["airport_fee"] + bd["water"]).round(2))
wbk = wb.create_sheet("Cost breakdown")
wbk.sheet_view.showGridLines = False
wbk["A1"] = "Full component breakdown — every country and scenario, USD-2024 per tonne of e-SAF"
wbk["A1"].font = Font(name=ARIAL, bold=True, size=14)
wbk["A2"] = ("Recomputed from the model's own cost functions at each run's chosen configuration and reconciled exactly against the run totals "
             "(column 'check_diff' = components minus the model's delivered cost; 0.00 everywhere). Each process line includes its share of "
             "balance-of-plant, owner's costs, contingency and fixed O&M.")
wbk["A2"].font = Font(name=ARIAL, size=9, italic=True, color="595959")

bd_headers = {
    "country": ("Country", 8), "scenario": ("Scenario", 22), "plant_kt": ("Plant (kt/y)", 10),
    "strategy": ("Strategy", 15), "carbon": ("Carbon", 9),
    "wind": ("Wind", 9), "solar_pv": ("Solar PV", 9), "battery": ("Battery", 8),
    "seasonal_h2_store": ("Seasonal H2 store", 10), "fuel_tank_store": ("Fuel tank store", 10),
    "electrolyser_flex_oversize": ("Electrolyser flex oversize", 11),
    "electrolyser_incl_stacks": ("Electrolyser (incl. stacks)", 11), "dac": ("DAC", 9),
    "ft_rwgs_upgrading": ("FT + RWGS + upgrading", 11), "heat_pump": ("Heat pump", 8),
    "compressors": ("Compressors", 9), "buffers_co2_h2": ("Buffers (CO2+H2)", 9),
    "fixed_site_services": ("Fixed site services", 10),
    "purchased_biogenic_co2": ("Purchased biogenic CO2", 11), "fuel_transport": ("Fuel transport", 9),
    "airport_fee": ("Airport fee", 8), "water": ("Water", 8),
    "small_lines_subtotal": ("Small lines subtotal (transport+fee+water)", 12),
    "total": ("TOTAL", 10), "model_delivered": ("Model delivered", 10), "check_diff": ("check_diff", 8),
}
HB = 4
for j, col in enumerate(bd.columns, start=1):
    h, w = bd_headers[col]
    c = wbk.cell(row=HB, column=j, value=h)
    c.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=8)
    c.fill = HDR_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    wbk.column_dimensions[get_column_letter(j)].width = w
wbk.row_dimensions[HB].height = 44
STRAT_S = {"flex_fuel_store": "flex + fuel tanks", "steady_h2_store": "steady + H2 cavern"}
CARB_S = {"market": "biogenic", "dac": "DAC"}
for i, (_, r0) in enumerate(bd.iterrows()):
    rr = HB + 1 + i
    for j, col in enumerate(bd.columns, start=1):
        v = r0[col]
        if col == "strategy":
            v = STRAT_S.get(v, v)
        elif col == "carbon":
            v = CARB_S.get(v, v)
        c = wbk.cell(row=rr, column=j, value=v)
        c.font = Font(name=ARIAL, size=8.5)
        c.border = BORDER
        if i % 2 == 1:
            c.fill = BAND
        if col == "plant_kt":
            c.number_format = "#,##0.0"
        elif col not in ("country", "scenario", "strategy", "carbon"):
            c.number_format = "#,##0.00" if col in ("airport_fee", "check_diff") else "#,##0"
wbk.freeze_panes = f"F{HB+1}"
wbk.auto_filter.ref = f"A{HB}:{get_column_letter(len(bd.columns))}{HB + len(bd)}"

# --------------------------------------------------------- Column definitions
wd = wb.create_sheet("Column definitions")
wd.sheet_view.showGridLines = False
wd["A1"] = "Column definitions and scenario notes"
wd["A1"].font = Font(name=ARIAL, bold=True, size=14)

defs = [
    ("Country / Country name / Hub airport / IATA", "The model places one plant per country and delivers to a single national hub airport (the busiest by fuel demand)."),
    ("Hub fuel 2024 (kt)", "Estimated jet fuel uplifted at the hub in 2024, thousand tonnes. Built from Eurostat national kerosene deliveries x the hub's share of national traffic, cross-checked against ICCT/ODI Airport Tracker departing-flight fuel burn (+/-10-20%). EASA's ReFuelEU report publishes no per-airport fuel."),
    ("e-SAF share 2030 / 2050", "The regulated synthetic-fuel (RFNBO) share of all fuel supplied at the airport: ReFuelEU Annex I minimums (1.2% average 2030-31; 35% in 2050); UK SAF Mandate PtL sub-obligation for Heathrow (~0.5% in 2030, ~3.5% from 2040, held flat); Switzerland applies ReFuelEU at Zurich from 2026."),
    ("Plant size 2030 / 2050 (kt/y)", "Plant capacity = hub fuel 2024 x fuel growth (EASA EAER 2025 base: +2% to 2030, ~0% to 2050 - traffic growth largely offset by efficiency) x the e-SAF share. Thousand tonnes of e-SAF per year."),
    ("Cost 2030 (USD/t)", "Delivered cost of one tonne of e-SAF at the hub airport for the 2030-mandate plant at 2030 technology costs: renewables + battery + seasonal storage + process plant + purchased CO2 + fuel transport + airport receiving fee + water, annualised at the country WACC. Real USD-2024. NOT a market price: no margin, taxes, subsidies or certificate value."),
    ("Cost 2030 (EUR/t)", "The same figure converted at the editable USD->EUR rate on the Results sheet (2024 ECB annual average 0.9239)."),
    ("Strategy 2030 / 2050", "The cheaper of two operating strategies chosen by the optimiser: 'flexible + fuel tanks' = synthesis train oversized, follows the renewable seasons, finished fuel stored in tanks; 'steady + H2 cavern' = constant-output plant with a seasonal hydrogen store priced by national geology (salt cavern vs lined rock)."),
    ("Carbon 2030 / 2050", "The cheaper RFNBO-eligible carbon source: 'biogenic (purchased)' = CO2 bought from bioenergy/pulp/waste point sources within trucking range (Del. Reg. (EU) 2023/1185, Annex pt. 10(c)); 'DAC (on-site)' = direct air capture on the plant's own meter (pt. 10(b)). Fossil industrial CO2 is excluded (eligibility ends 2036/2041, inside the plant's life)."),
    ("Bio-CO2 price 2030 (USD/t CO2)", "Delivered cost of the purchased biogenic CO2 where chosen: capture + liquefaction + purity polishing + transfer + trucking over the mean collection haul implied by national source density. Blank where DAC is chosen."),
    ("Cost 2050 @2030 tech (USD/t)", "2050 mandate demand served with 2030 technology costs - isolates the pure scale/demand effect."),
    ("Cost 2050 @2050 tech (USD/t and EUR/t)", "2050 mandate demand with published 2050 learning values for electrolyser, DAC, FT, PV, wind and battery (DEA / Fasihi / Seymour / NREL 2050 rows); financing and logistics stay at 2030 values. The headline 2050 number."),
    ("Reference 740 kt (USD/t)", "Same model at the fixed 740 kt/y reference size used in earlier analyses (2030 technology) - the comparability anchor across model versions."),
    ("WACC (real)", "Country-specific real pre-tax weighted average cost of capital applied to all capex (AURES II / IRENA surveys; UK DESNZ; CH SFOE; floor 4%, cap 10%). The process plant carries a +1pp premium."),
    ("H2 storage geology", "'salt' = onshore salt-cavern potential (Caglayan et al. 2020) - cheap seasonal H2 storage; 'rock' = no salt deposits, lined rock cavern assumed at ~2x the cost. Only matters where the steady + H2 cavern strategy is chosen."),
    ("— Import case sheet —", "Delivered cost of IMPORTED e-SAF at every EU-27 hub vs producing domestically, 2030 and 2050. Exporters (SAU/ARE/MAR) site at their best renewables near a designated export terminal (candidates tried; cheapest kept — UAE via Jebel Ali, Saudi Arabia via Yanbu King Fahd Port, Morocco via Ad Dakhla) and size ONE production system to 50% of the EU-27 synthetic mandate (197 kt/y 2030, 5.64 Mt/y 2050 from EASA's 32.2 Mt Union uplift) - the import share capped at half the market, mirroring the REPowerEU renewable-hydrogen balance - so every EU country faces the same production cost and differs only in the sea lane, its EU ETS cost and the inland legs. Sea routing, emission anchors (9 g CO2/t-nm 2030, 5 g 2050) and the 50% EU / 50% origin carbon split are ported unchanged from the user's ammonia trade model; kerosene product-tanker voyage costs (MR 40 kt in 2030, LR2 90 kt in 2050) replace the ammonia cost curve. Import port per country = nearest energy-capable WPI port to the hub airport (national port for coastal countries; nearest foreign port for landlocked). TOTAL and 'import vs domestic' are live formulas over the component columns."),
    ("— Ticket impact sheet —", "Translates the e-SAF mandate into end-consumer ticket increases for 2030 and 2050, vs the same flight burning 100% fossil jet. The e-SAF price airlines face is a live 50/50 blend of domestic production (EU-27 demand-weighted) and imports from the cheapest exporter (Import case sheet; import share editable, REPowerEU-style balance). Method: synthetic share x (blended e-SAF cost - fossil jet price) x fuel burn per passenger x cabin seat-area factor, divided by a representative fare; plus the operator-level view (fuel-bill % increase x fuel share of operating cost = uniform pass-through on every ticket). All drivers are editable input cells; a net-of-ETS column credits the EU ETS allowances the displaced fossil fuel would have needed on intra-EEA routes."),
    ("— Cost breakdown sheet —", "One row per country x scenario (128 rows), every component in USD-2024 per tonne. Lines: Wind / Solar PV / Battery / Seasonal H2 store / Fuel tank store / Electrolyser flex oversize (extra electrolyser capacity some steady designs buy for diurnal flexibility) / Electrolyser incl. stack replacement / DAC / FT+RWGS+upgrading / Heat pump / Compressors / Buffers / Fixed site services / Purchased biogenic CO2 / Fuel transport / Airport fee / Water. 'Small lines subtotal' collects transport + airport fee + water - real but small (typically 50-90 USD/t, 1-2.5% of total; see the zoom panel in Figure 3). Purchased biogenic CO2 is 500-820 USD/t where chosen in 2030 and zero in the 2050 runs (all hubs flip to on-site DAC, whose cost then sits in the DAC and renewables lines). 'check_diff' proves the reconciliation: components minus the model's delivered total, 0.00 everywhere."),
]
row = 3
wd.column_dimensions["A"].width = 34
wd.column_dimensions["B"].width = 120
for h, d in defs:
    a = wd.cell(row=row, column=1, value=h)
    a.font = Font(name=ARIAL, bold=True, size=9)
    a.alignment = Alignment(vertical="top", wrap_text=True)
    b = wd.cell(row=row, column=2, value=d)
    b.font = Font(name=ARIAL, size=9)
    b.alignment = Alignment(vertical="top", wrap_text=True)
    wd.row_dimensions[row].height = max(26, 13 * (len(d) // 118 + 1))
    row += 1

row += 1
wd.cell(row=row, column=1, value="Reader's caveats").font = Font(name=ARIAL, bold=True, size=11)
row += 1
cavs = [
    "Base-case values: several input ranges are right-skewed (FOAK evidence), so these deterministic numbers behave like a ~p10 of the model's own uncertainty; the sampled p50 runs ~25-35% higher. Quote the p10-p50-p90 band for headline claims.",
    "Resource layer: these runs use the offline heuristic solar/wind layer on a 100 km grid for reproducibility; NASA POWER / PVGIS re-runs sharpen site differentiation but do not change the input economics.",
    "Micro plants: 2030 mandate volumes at small hubs (SVN 0.2 kt/y, SVK, EST, LTU, HRV, LVA, BGR under ~2 kt/y) are far below minimum economic plant scale; their costs are reported honestly but read as 'import or aggregate', not as investable projects.",
    "Comparators: EASA 2025 reference price for synthetic fuel from atmospheric CO2 is EUR 7,815-9,525/t; Project SkyPower puts 2030 European first plants at EUR 5,000-8,000/t and the long-term floor at EUR 3,000-4,000/t.",
    "Full provenance: every input's source, currency and vintage is in literature_inputs.json / country_finance.json / country_biogenic_co2.json / policy_demand.json; model mechanics in NUMBERS_REVIEW.md and SCALE_AND_CARBON_REVIEW.md.",
]
for cv in cavs:
    c = wd.cell(row=row, column=2, value=cv)
    c.font = Font(name=ARIAL, size=9)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    wd.row_dimensions[row].height = max(26, 13 * (len(cv) // 118 + 1))
    row += 1

# --------------------------------------------------------- Ticket impact
ti = wb.create_sheet("Ticket impact", 1)   # right after Results
ti.sheet_view.showGridLines = False
YEL = PatternFill("solid", fgColor="FFFF00")

def lab(cell, v, size=9, bold=False, italic=False, color="000000", wrap=False):
    c = ti[cell]
    c.value = v
    c.font = Font(name=ARIAL, size=size, bold=bold, italic=italic, color=color)
    if wrap:
        c.alignment = Alignment(vertical="top", wrap_text=True)
    return c

def inp(cell, v, fmt=None):
    c = ti[cell]
    c.value = v
    c.font = INPUT_FONT
    c.fill = YEL
    if fmt:
        c.number_format = fmt
    return c

def frm(cell, f, fmt=None):
    c = ti[cell]
    c.value = f
    c.font = Font(name=ARIAL, size=9)
    if fmt:
        c.number_format = fmt
    return c

for col, w in [("A", 40), ("B", 11), ("C", 13), ("D", 11), ("E", 12), ("F", 12),
               ("G", 11), ("H", 13), ("I", 12), ("J", 11), ("K", 13)]:
    ti.column_dimensions[col].width = w

lab("A1", "What the e-SAF mandate adds to an airline ticket — EU average, 2030 and 2050", 14, bold=True)
lab("A2", "Extra cost per passenger = synthetic share x (e-SAF delivered cost - fossil jet price) x fuel burn per passenger x cabin factor; "
          "ticket increase (%) = extra cost / representative fare. Airline passes the full cost to the consumer (linear pass-through).", 9, italic=True, color="595959")
lab("A3", "Counterfactual: the same flight burning 100% fossil jet at the price below. Bio-SAF obligations exist in both worlds and cancel out - "
          "only the synthetic (e-SAF) share is priced here. Blue on yellow = editable inputs; everything recalculates.", 9, italic=True, color="595959")

lab("A5", "Shared inputs", 11, bold=True)
shared = [
    (6,  "Fossil jet fuel price (USD-2024 / t)", 780, "#,##0",
     "IATA Global Outlook Dec-2025: 2024 average 99 USD/bbl x 7.88 bbl/t; held flat in real terms (long-run range ~650-950)."),
    (7,  "Fuel share of airline operating cost", 0.288, "0.0%",
     "IATA, 2024 actual (28.8%; forecast ~26-27% for 2025-26)."),
    (8,  "e-SAF share of all fuel, 2030", 0.012, "0.0%",
     "ReFuelEU Annex I synthetic sub-mandate, average 2030-31. UK PtL sub-obligation is 0.5% - scale 2030 results by ~0.42 for UK operators."),
    (9,  "e-SAF share of all fuel, 2050", 0.35, "0%",
     "ReFuelEU Annex I, 2050. UK: ~3.5% from 2040 held flat - scale 2050 results by ~0.10 for UK operators."),
    (10, "Domestic e-SAF cost 2030 (USD / t)", 7056, "#,##0",
     "This model: EU-27 demand-weighted mean across national hubs, 2030 mandate plants at 2030 technology (p10-p90: 5,898-8,047)."),
    (11, "Domestic e-SAF cost 2050 (USD / t)", 3207, "#,##0",
     "This model: 2050 mandate plants at 2050 technology (p10-p90: 2,821-3,603)."),
    (12, "Imported e-SAF cost 2030 (USD / t)", 7076, "#,##0",
     "Import case sheet: demand-weighted best-exporter delivered cost, exporters sized to 50% of EU demand (p10-p90: 7,051-7,106)."),
    (13, "Imported e-SAF cost 2050 (USD / t)", 3041, "#,##0",
     "Import case sheet (p10-p90: 3,020-3,069)."),
    (14, "Import share of e-SAF supply", 0.50, "0%",
     "REPowerEU-style balance: half the mandate met by imports (cf. 10 Mt domestic + 10 Mt imported renewable H2 by 2030)."),
]
for r, label_, v, fmt, src in shared:
    lab(f"A{r}", label_)
    inp(f"B{r}", v, fmt)
    lab(f"C{r}", src, 8, italic=True, color="595959")

lab("A15", "e-SAF cost used, 2030 (USD / t)")
frm("B15", "=(1-$B$14)*B10+$B$14*B12", "#,##0")
lab("C15", "blend feeding every calculation below; edit the import share in B14", 8, italic=True, color="595959")
lab("A16", "e-SAF cost used, 2050 (USD / t)")
frm("B16", "=(1-$B$14)*B11+$B$14*B13", "#,##0")

shared2 = [
    (17, "EU ETS allowance price 2030 (USD / t CO2)", 141, "#,##0",
     "~EUR 130 x 1.0824 USD/EUR; analyst forecasts cluster EUR 126-149 by 2030 (BNEF and market surveys)."),
    (18, "EU ETS allowance price 2050 (USD / t CO2)", 271, "#,##0",
     "~EUR 250; long-run projections EUR 200-400. Used only in the net-of-ETS columns."),
    (19, "Fossil jet combustion CO2 (t CO2 / t fuel)", 3.16, "0.00", "Standard jet A-1 emission factor."),
    (20, "Fleet fuel-burn factor 2030 (vs 2024)", 1.00, "0.00",
     "1.00 = conservative (no efficiency gain). ~1%/y fleet renewal would give ~0.94."),
    (21, "Fleet fuel-burn factor 2050 (vs 2024)", 1.00, "0.00",
     "~0.77 with 1%/y efficiency to 2050; scales the per-passenger premium down proportionally."),
]
for r, label_, v, fmt, src in shared2:
    lab(f"A{r}", label_)
    inp(f"B{r}", v, fmt)
    lab(f"C{r}", src, 8, italic=True, color="595959")

lab("A23", "e-SAF premium over fossil jet, 2030 (USD / t)")
frm("B23", "=B15-B6", "#,##0")
lab("A24", "e-SAF premium over fossil jet, 2050 (USD / t)")
frm("B24", "=B16-B6", "#,##0")
lab("A25", "2030 premium net of EU ETS saved (USD / t)")
frm("B25", "=B23-B19*B17", "#,##0")
lab("C25", "displaced fossil fuel would have needed allowances (intra-EEA departures only)", 8, italic=True, color="595959")
lab("A26", "2050 premium net of EU ETS saved (USD / t)")
frm("B26", "=B24-B19*B18", "#,##0")

lab("A28", "Per-segment result (one-way trip)", 11, bold=True)
seg_hdr = ["Route / cabin", "Sector (km)", "Fuel kg / economy pax", "Cabin fuel factor",
           "One-way fare (USD)", "Extra USD/pax 2030", "Ticket +% 2030", "+% 2030 net of ETS*",
           "Extra USD/pax 2050", "Ticket +% 2050", "+% 2050 net of ETS*"]
for j, h in enumerate(seg_hdr, start=1):
    c = ti.cell(row=29, column=j, value=h)
    c.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=8)
    c.fill = HDR_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
ti.row_dimensions[29].height = 40

segs = [
    ("Short haul - Economy", 1250, 30, 1.0, 110),
    ("Short haul - Business", 1250, 30, 1.5, 330),
    ("Long haul - Economy", 7000, 182, 1.0, 450),
    ("Long haul - Business", 7000, 182, 2.9, 2200),
]
for i, (name, km, kg, fac, fare) in enumerate(segs):
    r = 30 + i
    lab(f"A{r}", name)
    lab(f"B{r}", km).number_format = "#,##0"
    inp(f"C{r}", kg, "#,##0")
    inp(f"D{r}", fac, "0.0")
    inp(f"E{r}", fare, "#,##0")
    frm(f"F{r}", f"=$B$8*$B$23*C{r}*D{r}*$B$20/1000", "0.00")
    frm(f"G{r}", f"=F{r}/E{r}", "0.0%")
    frm(f"I{r}", f"=$B$9*$B$24*C{r}*D{r}*$B$21/1000", "0.00")
    frm(f"J{r}", f"=I{r}/E{r}", "0.0%")
    if i < 2:  # intra-EEA short haul: EU ETS applies
        frm(f"H{r}", f"=$B$8*$B$25*C{r}*D{r}*$B$20/1000/E{r}", "0.0%")
        frm(f"K{r}", f"=$B$9*$B$26*C{r}*D{r}*$B$21/1000/E{r}", "0.0%")
    else:
        lab(f"H{r}", "outside EU ETS", 8, italic=True, color="595959")
        lab(f"K{r}", "outside EU ETS", 8, italic=True, color="595959")
    for col in "ABCDEFGHIJK":
        ti[f"{col}{r}"].border = BORDER

lab("A36", "Operator-level view (whole network)", 11, bold=True)
lab("A37", "Increase in the airline's total fuel bill")
frm("G37", "=B8*B23/B6", "0.0%")
frm("J37", "=B9*B24/B6", "0.0%")
lab("A38", "Uniform pass-through on every ticket (fuel bill % x fuel share of opex)")
frm("G38", "=B7*G37", "0.0%")
frm("H38", "=B7*B8*B25/B6", "0.0%")
frm("J38", "=B7*J37", "0.0%")
frm("K38", "=B7*B9*B26/B6", "0.0%")
lab("C38", "same % on every ticket if the airline spreads the cost pro-rata over operating cost", 8, italic=True, color="595959")

notes = [
    "e-SAF supply blend: each hub covers half its synthetic mandate with national production and half with imports from the cheapest exporter (Import case sheet; exporters capped at 50% of EU demand, mirroring the REPowerEU renewable-hydrogen balance). The blend cell B15/B16 is live - set B14 to 0% for the domestic-only view or 100% for import-only.",
    "* Net of ETS: credits the EU ETS allowances the displaced fossil fuel would have needed. Applies to intra-EEA departures (short haul here); long-haul extra-EU routes sit outside the EU ETS scope and CORSIA unit prices are negligible by comparison.",
    "Fuel per passenger: ~24 g/pax-km short haul and ~26 g/pax-km long haul (economy basis), consistent with UK DEFRA per-class flight factors and the ICCT transatlantic average of 34 pax-km/L. Cabin factors are DEFRA seat-area multipliers (business 1.5x short haul, 2.9x long haul).",
    "Fares are representative one-way values, deliberately editable: intra-EU economy ~110 USD (IATA global average one-way fare ~168 USD incl. ancillaries), euro-business ~3x economy, long-haul economy ~450 USD, long-haul business ~2,200 USD. The % result scales 1:1 with the fare you assume.",
    "Both views are consistent: the per-segment rows attribute fuel by seat area (business burns more, but its fare base is larger, so its % is lower); the operator row spreads the same total cost evenly. Reality sits between them, set by each airline's pricing power per cabin.",
    "The e-SAF costs are this model's delivered costs (no producer margin, no certificate value, no penalty avoidance). ReFuelEU's flexibility mechanism (2030-34) and mandate penalties would change airline strategy, not the underlying cost per tonne.",
]
r = 40
for n in notes:
    lab(f"A{r}", n, 8, italic=True, color="595959")
    ti.merge_cells(f"A{r}:K{r}")
    ti[f"A{r}"].alignment = Alignment(vertical="top", wrap_text=True)
    ti.row_dimensions[r].height = max(24, 11 * (len(n) // 150 + 1))
    r += 1

# --------------------------------------------------------- Import case
imp_path = Path(__file__).resolve().parent.parent / "import_case" / "import_delivered_costs.csv"
if imp_path.exists():
    imp = pd.read_csv(imp_path)
    prod = pd.read_csv(imp_path.parent / "export_production_raw.csv")
    ic = wb.create_sheet("Import case", 2)
    ic.sheet_view.showGridLines = False
    ic["A1"] = "Importing e-SAF from SAU / ARE / MAR vs producing domestically — every EU-27 hub, 2030 and 2050"
    ic["A1"].font = Font(name=ARIAL, bold=True, size=14)
    ic["A2"] = ("Exporters size one production system to 50% of the EU-27 synthetic mandate (197 kt/y 2030; 5.64 Mt/y 2050) - a REPowerEU-style import balance (cf. 10 Mt domestic + 10 Mt imported renewable H2) - "
                "at their best renewables near the chosen export terminal. Chain: production at terminal + sea freight + EU ETS on voyage emissions + import-port handling + inland to the hub airport + airport fee. All USD-2024 per tonne.")
    ic["A2"].font = Font(name=ARIAL, size=9, italic=True, color="595959")
    ic["A3"] = ("Sea routing, emission anchors and the 50% EU / 50% origin carbon treatment are ported unchanged from the ammonia trade model (AP_IRA_Model); "
                "kerosene product-tanker costs replace the ammonia curve. See import_case/IMPORT_CASE_README.md for provenance, validation and caveats.")
    ic["A3"].font = Font(name=ARIAL, size=9, italic=True, color="595959")

    ic["A5"] = "Export production (plant sized to the whole EU-27 mandate; DAC carbon; country WACC)"
    ic["A5"].font = Font(name=ARIAL, bold=True, size=11)
    ph = ["Exporter", "Year", "Export terminal", "Site lat", "Site lon", "Site→terminal (km)", "Inland mode",
          "Production at terminal (USD/t)", "WACC (real)", "Terminal candidates tried (USD/t)"]
    for j, hh in enumerate(ph, start=1):
        c = ic.cell(row=6, column=j, value=hh)
        c.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=8)
        c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ic.row_dimensions[6].height = 30
    for i, (_, r0) in enumerate(prod.iterrows()):
        rr = 7 + i
        vals = [r0["exporter"], int(r0["year"]), r0["export_port"], round(r0["best_lat"], 2), round(r0["best_lon"], 2),
                round(r0["site_to_terminal_km"]), r0["transport_mode"],
                round(r0["delivered_cost_usd_per_tonne_saf"] - 5.412 + 5.0), r0["wacc_used"], r0["terminals_tried"]]
        for j, v in enumerate(vals, start=1):
            c = ic.cell(row=rr, column=j, value=v)
            c.font = Font(name=ARIAL, size=9)
            c.border = BORDER
        ic.cell(row=rr, column=8).number_format = "#,##0"
        ic.cell(row=rr, column=9).number_format = "0.0%"

    HDR2 = 15
    ic[f"A{HDR2-1}"] = "Delivered import cost per hub (all exporters) — TOTAL and comparison are live formulas"
    ic[f"A{HDR2-1}"].font = Font(name=ARIAL, bold=True, size=11)
    heads = [("Country", 8), ("Year", 7), ("Exporter", 9), ("Export port", 14), ("Import port", 16),
             ("Route (nm)", 9), ("Via Suez", 8), ("Vessel", 7), ("Production at terminal", 11),
             ("Sea freight", 9), ("Ship CO2 (kg/t)", 9), ("EU ETS on voyage", 9), ("Import port handling", 9),
             ("Inland to airport", 9), ("Airport fee", 8), ("TOTAL import (USD/t)", 11),
             ("Domestic (USD/t)", 11), ("Import vs domestic", 10), ("Cheapest option", 12)]
    for j, (hh, w) in enumerate(heads, start=1):
        c = ic.cell(row=HDR2, column=j, value=hh)
        c.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=8)
        c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ic.column_dimensions[get_column_letter(j)].width = w
    ic.row_dimensions[HDR2].height = 40
    imp_sorted = imp.sort_values(["country", "year", "exporter"]).reset_index(drop=True)
    best_key = imp.loc[imp.groupby(["country", "year"])["import_delivered_usd_t"].idxmin()]
    best_set = {(r.country, r.year, r.exporter) for r in best_key.itertuples()}
    for i, (_, r0) in enumerate(imp_sorted.iterrows()):
        rr = HDR2 + 1 + i
        is_best = (r0["country"], r0["year"], r0["exporter"]) in best_set
        vals = [r0["country"], int(r0["year"]), r0["exporter"], r0["export_port"], r0["import_port"],
                r0["distance_nm"], "yes" if r0["uses_suez"] else "no", r0["vessel"],
                r0["production_usd_t"], r0["shipping_usd_t"], r0["ship_co2_kg_t"], r0["ets_usd_t"],
                r0["import_port_handling_usd_t"], r0["inland_usd_t"], r0["airport_fee_usd_t"],
                None, r0["domestic_usd_t"], None,
                ("BEST — beats domestic" if r0["import_delivered_usd_t"] < r0["domestic_usd_t"] else "BEST — domestic cheaper") if is_best else None]
        for j, v in enumerate(vals, start=1):
            c = ic.cell(row=rr, column=j, value=v)
            c.font = Font(name=ARIAL, size=8.5, bold=(is_best and j >= 16))
            c.border = BORDER
            if i % 2 == 1:
                c.fill = BAND
        ic.cell(row=rr, column=16, value=f"=SUM(I{rr}:J{rr})+SUM(L{rr}:O{rr})").font = Font(name=ARIAL, size=8.5, bold=is_best)
        ic.cell(row=rr, column=18, value=f"=IF(Q{rr}=\"\",\"\",P{rr}/Q{rr}-1)").font = Font(name=ARIAL, size=8.5, bold=is_best)
        for col, fmt in [(6, "#,##0"), (9, "#,##0"), (10, "#,##0.0"), (11, "#,##0.0"), (12, "0.00"),
                         (13, "0.0"), (14, "#,##0.0"), (15, "0.00"), (16, "#,##0"), (17, "#,##0"), (18, "+0.0%;-0.0%")]:
            ic.cell(row=rr, column=col).number_format = fmt
    last2 = HDR2 + len(imp_sorted)
    ic.freeze_panes = f"D{HDR2+1}"
    ic.auto_filter.ref = f"A{HDR2}:S{last2}"

    notes_ic = [
        "Import ports: nearest energy-capable WPI port to the hub airport (oil-terminal depth >= 9 m, oil-terminal flag, or Large/Medium liquid bulk). Coastal countries keep a national port (SVN Koper, MLT Valletta relaxed); landlocked use the nearest foreign energy port (AUT Trieste, CZE Szczecin, HUN+SVK Bakar, LUX Vlissingen).",
        "Sea freight: round-trip product-tanker voyage (charter + bunkers at 550 USD/t VLSFO + port costs + Suez toll 325k USD x2 transits where used) over the cargo parcel; year multipliers 1.00 (2030) -> 0.80 (2050); +/-30% band available in shipping_kerosene.py. Ship CO2 = distance x 9 g/t-nm (2030) or 5 g/t-nm (2050); EU ETS = 50% x 141 / 271 USD per t CO2 (origin carbon price 0).",
        "Production at terminal swaps the airport fee for 5 USD/t terminal handling; import chain adds 5 USD/t port handling, road inland at 0.092 USD/t-km x 1.35 routing factor, and the standard 5.41 USD/t airport fee.",
        "Ad Dakhla is a greenfield terminal (Morocco-administered Western Sahara; Morocco's planned southern green-fuel corridor) — its 'Very Small' WPI status reflects today's port, not the assumed build-out. Exporter ranking is a financing story: UAE 8.0% vs SAU 8.7% vs MAR 10.0% real pre-tax WACC; the whole sea leg incl. ETS is 10-72 USD/t (0.3-2.5% of delivered cost).",
        "No producer margin, certificate value or penalty avoidance on either side; RFNBO import eligibility assumed (RED-compliant production). 2050 exporter volumes (11.27 Mt/y) represent a national export industry (number-up scaling), not a single plant.",
    ]
    r = last2 + 2
    for n in notes_ic:
        c = ic.cell(row=r, column=1, value=n)
        c.font = Font(name=ARIAL, size=8, italic=True, color="595959")
        ic.merge_cells(f"A{r}:S{r}")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ic.row_dimensions[r].height = max(24, 11 * (len(n) // 160 + 1))
        r += 1

wb.save(OUT)
print("written", OUT)
